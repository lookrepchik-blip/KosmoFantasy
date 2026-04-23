from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


@dataclass(slots=True)
class Player:
    name: str
    position: str
    participant_name: str | None = None
    source_sheet: str | None = None
    team_name: str | None = None


class ExcelRepository:
    def __init__(self, path: Path):
        self.path = path

    def _find_header_row(self, ws, required_headers: set[str], max_scan: int = 10):
        for row in range(1, min(ws.max_row, max_scan) + 1):
            values = [str(ws.cell(row, col).value).strip().lower() if ws.cell(row, col).value is not None else "" for col in range(1, ws.max_column + 1)]
            if required_headers.issubset(set(values)):
                return row, values
        return None, []

    def load_participants(self) -> list[str]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path, data_only=True)

        for name in ["Participants", "Участники"]:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            header_row, header = self._find_header_row(ws, {"участник"})
            if not header_row:
                header_row, header = self._find_header_row(ws, {"participant"})
            if not header_row:
                continue
            idx = header.index("участник") + 1 if "участник" in header else header.index("participant") + 1
            out: list[str] = []
            for row in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(row, idx).value
                if val:
                    out.append(str(val).strip())
            if out:
                return out

        if "Lists" in wb.sheetnames:
            ws = wb["Lists"]
            values: list[str] = []
            for row in range(3, ws.max_row + 1):
                val = ws.cell(row, 1).value
                if not val:
                    continue
                sval = str(val).strip()
                if sval.lower().startswith("клуб") or sval.lower().startswith("clubs"):
                    break
                values.append(sval)
            if values:
                return values
        return []

    def load_global_pool(self) -> list[Player]:
        if not self.path.exists():
            return []
        wb = load_workbook(self.path, data_only=True)
        if "BeteraPool" not in wb.sheetnames:
            return []
        ws = wb["BeteraPool"]
        header_row, header = self._find_header_row(ws, {"игрок", "клуб", "позиция"}, max_scan=10)
        if not header_row:
            header_row, header = self._find_header_row(ws, {"player", "team", "position"}, max_scan=10)
        if not header_row:
            return []

        def idx_of(*names: str) -> int | None:
            for name in names:
                if name in header:
                    return header.index(name) + 1
            return None

        player_col = idx_of("игрок", "player", "player_name", "имя")
        team_col = idx_of("клуб", "team", "club")
        pos_col = idx_of("позиция", "position")
        active_col = idx_of("активен", "active")
        if not player_col:
            return []

        raw_players: list[Player] = []
        for row in range(header_row + 1, ws.max_row + 1):
            player_val = ws.cell(row, player_col).value
            if not player_val:
                continue
            active_ok = True
            if active_col:
                active_val = ws.cell(row, active_col).value
                if active_val is not None and str(active_val).strip().lower() in {"нет", "no", "false", "0"}:
                    active_ok = False
            if not active_ok:
                continue
            pos = str(ws.cell(row, pos_col).value).strip() if pos_col and ws.cell(row, pos_col).value else ""
            team = str(ws.cell(row, team_col).value).strip() if team_col and ws.cell(row, team_col).value else None
            raw_players.append(Player(str(player_val).strip(), pos, participant_name=None, source_sheet="BeteraPool", team_name=team))

        counts: dict[str, int] = {}
        for player in raw_players:
            counts[player.name] = counts.get(player.name, 0) + 1
        out: list[Player] = []
        for player in raw_players:
            name = player.name
            if counts.get(player.name, 0) > 1 and player.team_name:
                name = f"{player.name} [{player.team_name}]"
            out.append(Player(name, player.position, participant_name=None, source_sheet=player.source_sheet, team_name=player.team_name))
        return out

    def load_rules_text(self) -> str:
        if not self.path.exists():
            return "Правила не найдены: Excel-файл отсутствует."
        wb = load_workbook(self.path, data_only=True)
        sheet_name = None
        for candidate in ["Rules", "Правила"]:
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break
        if not sheet_name:
            return "Правила не найдены в Excel."
        ws = wb[sheet_name]
        lines: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if not vals:
                if lines and lines[-1] != "":
                    lines.append("")
                continue
            if len(vals) >= 3 and vals[1] and vals[2]:
                lines.append(f"{vals[0]} — {vals[1]}: {vals[2]}")
            else:
                lines.append(" | ".join(vals))
        text = "\n".join(lines).strip()
        return text or "Правила в Excel пустые."

    def load_rosters(self) -> dict[str, list[Player]]:
        global_pool = self.load_global_pool()
        participants = self.load_participants()
        if global_pool and participants:
            return {participant: list(global_pool) for participant in participants}
        if not self.path.exists():
            return {}
        wb = load_workbook(self.path, data_only=True)
        result: dict[str, list[Player]] = {}
        for sheet_name in ["Rosters_Reg", "Rosters_PO", "Составы_Регулярка", "Составы_Плейофф", "Lineups", "Rosters"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header_row, header = self._find_header_row(ws, {"участник", "позиция", "игрок"})
            if not header_row:
                header_row, header = self._find_header_row(ws, {"participant", "position", "player"})
            if not header_row:
                continue

            def idx_of(*names: str) -> int | None:
                for name in names:
                    if name in header:
                        return header.index(name) + 1
                return None

            part_col = idx_of("participant", "participant_name", "участник")
            player_col = idx_of("player", "player_name", "игрок", "имя")
            pos_col = idx_of("position", "позиция")
            if not part_col or not player_col:
                continue
            for row in range(header_row + 1, ws.max_row + 1):
                part_val = ws.cell(row, part_col).value
                player_val = ws.cell(row, player_col).value
                if not part_val or not player_val:
                    continue
                part = str(part_val).strip()
                player = str(player_val).strip()
                pos = str(ws.cell(row, pos_col).value).strip() if pos_col and ws.cell(row, pos_col).value else ""
                result.setdefault(part, []).append(Player(player, pos, participant_name=part, source_sheet=sheet_name))
        return result

    def export_bot_data(self, users: Iterable, lineups: Iterable, scores: Iterable, leaderboard: Iterable, output_path: Path | None = None, transfers: Iterable | None = None) -> Path:
        output_path = output_path or self.path.with_name(self.path.stem + "_bot_export.xlsx")
        if self.path.exists():
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            default = wb.active
            wb.remove(default)

        for sheet_name in ["BotUsers", "BotLineups", "BotScores", "BotLeaderboard", "BotTransfers"]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        ws = wb.create_sheet("BotUsers")
        ws.append(["telegram_id", "username", "full_name", "participant_name", "is_admin", "created_at", "updated_at"])
        for row in users:
            ws.append([row["telegram_id"], row["username"], row["full_name"], row["participant_name"], row["is_admin"], row["created_at"], row["updated_at"]])

        ws = wb.create_sheet("BotLineups")
        ws.append(["participant_name", "stage", "round_no", "player_name", "position", "submitted_by", "submitted_at"])
        for row in lineups:
            ws.append([row["participant_name"], row["stage"], row["round_no"], row["player_name"], row["position"], row["submitted_by"], row["submitted_at"]])

        ws = wb.create_sheet("BotScores")
        ws.append(["participant_name", "stage", "round_no", "player_name", "position", "goals", "assists", "goals_allowed", "shutout", "manual_adj", "counted", "points", "created_by", "created_at", "updated_at"])
        for row in scores:
            ws.append([row["participant_name"], row["stage"], row["round_no"], row["player_name"], row["position"], row["goals"], row["assists"], row["goals_allowed"], row["shutout"], row["manual_adj"], row["counted"], row["points"], row["created_by"], row["created_at"], row["updated_at"]])

        ws = wb.create_sheet("BotLeaderboard")
        ws.append(["participant_name", "total_points", "entries"])
        for row in leaderboard:
            ws.append([row["participant_name"], row["total_points"], row["entries"]])

        ws = wb.create_sheet("BotTransfers")
        ws.append(["participant_name", "stage", "round_no", "quota_kind", "player_out", "player_in", "created_by", "created_at"])
        for row in transfers or []:
            ws.append([row["participant_name"], row["stage"], row["round_no"], row["quota_kind"], row["player_out"], row["player_in"], row["created_by"], row["created_at"]])

        wb.save(output_path)
        return output_path
